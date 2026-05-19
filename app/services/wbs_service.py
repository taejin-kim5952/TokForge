"""WBS 엑셀 양식 분석 + AI 자동 채우기 서비스.

흐름:
  1. analyze_template()  — .xlsx 업로드 → 양식 구조(컬럼/헤더 위치) 파악
  2. fill_template()     — LLM 추출 or 직접 입력 항목 → 양식에 삽입 → .xlsx 반환
"""

import json
import logging
import re
import uuid
from pathlib import Path

import httpx
import openpyxl
from openpyxl.utils import get_column_letter

from app.config import DEFAULT_MODEL, OLLAMA_BASE_URL

logger = logging.getLogger(__name__)

TEMPLATE_DIR = Path(__file__).parent.parent.parent / "storage" / "wbs_templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


# ──────────────────────────────────────────────
# 양식 분석
# ──────────────────────────────────────────────

def analyze_template(file_bytes: bytes, filename: str) -> dict:
    """업로드된 .xlsx 파일을 분석해 헤더 구조를 반환하고 파일을 임시 저장합니다.

    Returns:
        {
            "template_id": str,
            "filename": str,
            "sheet_name": str,
            "header_row": int,         # 1-based
            "data_start_row": int,     # 헤더 다음 첫 데이터 행 (1-based)
            "columns": [
                {"name": str, "col_letter": str, "col_index": int}
            ],
            "sample_rows": [[...], ...]  # 기존 데이터 최대 3행
        }
    """
    template_id = uuid.uuid4().hex
    template_path = TEMPLATE_DIR / f"{template_id}.xlsx"
    template_path.write_bytes(file_bytes)

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active

    # 헤더 행 탐색: 비어있지 않은 셀이 2개 이상 있는 첫 번째 행
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [v for v in row if v is not None and str(v).strip()]
        if len(non_empty) >= 2:
            header_row = row_idx
            break

    # 헤더 컬럼 수집
    columns = []
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = cell.value
        if val is not None and str(val).strip():
            columns.append({
                "name": str(val).strip(),
                "col_letter": get_column_letter(col_idx),
                "col_index": col_idx,
            })

    # 샘플 데이터 (헤더 다음 최대 3행)
    data_start_row = header_row + 1
    sample_rows = []
    for row in ws.iter_rows(
        min_row=data_start_row,
        max_row=data_start_row + 2,
        values_only=True,
    ):
        row_vals = [str(v) if v is not None else "" for v in row]
        if any(v for v in row_vals):
            sample_rows.append(row_vals)

    wb.close()

    return {
        "template_id": template_id,
        "filename": filename,
        "sheet_name": ws.title,
        "header_row": header_row,
        "data_start_row": data_start_row,
        "columns": columns,
        "sample_rows": sample_rows,
    }


# ──────────────────────────────────────────────
# LLM으로 대화 → WBS 항목 추출
# ──────────────────────────────────────────────

async def extract_items_from_conversation(
    conversation: str,
    columns: list[dict],
) -> list[dict]:
    """대화 내용을 LLM에 보내 WBS 항목 JSON 배열을 추출합니다."""
    col_names = [c["name"] for c in columns]
    prompt = f"""아래 대화에서 WBS 태스크 항목을 추출해 JSON 배열로만 답하세요.
다른 설명 없이 JSON만 출력하세요.

컬럼 목록: {col_names}
형식: [{{"컬럼명": "값", ...}}, ...]

비어있는 컬럼은 빈 문자열("")로 채우세요.
날짜는 YYYY-MM-DD 형식으로 변환하세요.

대화:
{conversation}

JSON:"""

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                f"{OLLAMA_BASE_URL}/api/chat",
                json={
                    "model": DEFAULT_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "stream": False,
                },
            )
            resp.raise_for_status()
            content = resp.json()["message"]["content"]

        # JSON 파싱 — 마크다운 코드 블록 제거 후 시도
        cleaned = re.sub(r"```(?:json)?", "", content).strip().rstrip("```").strip()
        items = json.loads(cleaned)
        if isinstance(items, dict):
            items = [items]
        return items if isinstance(items, list) else []

    except Exception as e:
        logger.error("WBS 항목 추출 실패: %s", e)
        return []


# ──────────────────────────────────────────────
# 양식에 항목 삽입 → 완성 파일 반환
# ──────────────────────────────────────────────

def fill_template(template_id: str, items: list[dict], header_row: int, columns: list[dict]) -> bytes:
    """items를 원본 양식에 삽입하고 완성된 .xlsx bytes를 반환합니다.

    - 원본 스타일(색상, 폰트, 테두리, 병합) 유지
    - 헤더 다음 행부터 순서대로 작성
    """
    template_path = TEMPLATE_DIR / f"{template_id}.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"양식 파일을 찾을 수 없습니다: {template_id}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 컬럼명 → col_letter 매핑
    col_map = {c["name"]: c["col_letter"] for c in columns}

    data_start_row = header_row + 1

    for row_offset, item in enumerate(items):
        target_row = data_start_row + row_offset
        for col_name, value in item.items():
            col_letter = col_map.get(col_name)
            if col_letter:
                ws[f"{col_letter}{target_row}"] = value if value != "" else None

    # 메모리에 저장 후 bytes 반환
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.read()


def delete_template(template_id: str) -> None:
    """임시 저장된 양식 파일 삭제."""
    path = TEMPLATE_DIR / f"{template_id}.xlsx"
    path.unlink(missing_ok=True)
